from openpyxl import load_workbook
import random
workbook = load_workbook(filename="bombs.xlsx")
workbook.sheetnames
sheet = workbook.active

list = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z"]
bomb = []

for y in range(26):
    for z in range(1,26):
        sheet[list[y]+str(z)] = " "
        
a = 0
for i in range(100):
    x = random.randint(1,26)
    if a < 26 :
        sheet[list[a]+str(x)] = "Bomb"
        bomb.append(list[a]+str(x))
        a += 1
    else:
        a = 0
print(bomb)

workbook.save(filename="bombs.xlsx")

for u in range(676):
    j = (input("Number : "))
    if j in bomb:
        print("Game over")
        break
    else:
        print("You are a survivor")
